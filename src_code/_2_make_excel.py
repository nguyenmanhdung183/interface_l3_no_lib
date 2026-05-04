import re
import json
import os
from _1_prepare_struct import *

# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from openpyxl.utils import get_column_letter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

DATA_EXCEL_PATH = os.path.join(BASE_DIR, "..", "data", "data.xlsx")

HANDLED_STRUCTS = set()  # để tránh lặp lại khi struct được dùng nhiều lần

#import pandas as pd



    
def format_excel(ws):
    # Căn giữa
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Tự động điều chỉnh độ rộng cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width    


# def map_to_c_type(type_str):
#     for c_type, aliases in PRIMITIVE_LIST.items():
#         if type_str in aliases:
#             return c_type
#     return type_str  # nếu không phải primitive → giữ nguyên (struct / enum)

def get_data_type(type_str):

    for key, values in PRIMITIVE_LIST.items():
        if type_str in values:
            return {
                "kind": "PRIMITIVE",
                "detail": key
            }

    return {
        "kind": "STRUCT",
        "detail": None
    }
        
def parse_fields_with_meta(struct_body):
    """
    Phiên bản nâng cấp của parse_fields:
    - Lấy meta: present, bitmask, min, max, description
    - Nhận biết array
    """
    fields = []
    for line in struct_body.split("\n"):
        line = line.strip()
        if not line:
            continue

        comment_matches = re.findall(r"/\*\^(.*?)\^\*/", line)
        metadata = {}

        if comment_matches:
            # annotation chính
            parts = [x.strip() for x in comment_matches[0].split(",")]
            
            if len(parts) >= 1:
                metadata["present"] = parts[0]
            if len(parts) >= 2:
                metadata["bitmask"] = parts[1]
            if len(parts) >= 3:
                metadata["range_check"] = parts[2]
                    # --- xử lý OCTET_STRING riêng ---
            if metadata.get("range_check") == "OCTET_STRING":
                if len(parts) >= 4:
                    metadata["description"] = parts[3]
            if len(parts) >= 4:
                metadata["min"] = parts[3]
            if len(parts) >= 5:
                metadata["max"] = parts[4]

            # annotation thứ 2 → dependency
            if len(comment_matches) >= 2:
                metadata["length_ref"] = comment_matches[1].strip()
        
        # # parse description // ...
        # desc_match = re.search(r"//(.*)", line)
        # if desc_match:
        #     metadata["description"] = desc_match.group(1).strip()
        
        if metadata.get("range_check") == "OCTET_STRING":
            metadata["description"] = parts[3] if len(parts) >= 4 else ""
            metadata["min"] = ""
            metadata["max"] = ""

        # remove comments
        line = re.sub(r"/\*\*.*?\*/", "", line)
        line = re.sub(r"//.*", "", line)

        # parse field type, name, array
        match = re.match(r"([\w_]+)\s+([\w_]+)(\[(.*?)\])?", line)
        if match:
            fields.append({
                "type": match.group(1),
                "name": match.group(2),
                "data_type": get_data_type(match.group(1))["kind"],
                "key" : get_data_type(match.group(1))["detail"],
                "is_array": match.group(3) is not None,
                "array_size": match.group(4) if match.group(4) else None,
                "meta": metadata
            })
    return fields

# ---------------- Build tree nâng cấp ----------------
def build_tree_with_meta(struct_name, structs, visited=None):
    """
    Build tree nested struct, giữ meta, check primitive
    """
    if visited is None:
        visited = set()
    if struct_name in visited:
        return {"type": struct_name, "fields": []}
    visited.add(struct_name)

    if struct_name not in structs:
        return {"type": struct_name, "fields": []}

    tree = {"type": struct_name, "fields": []}
    fields = parse_fields_with_meta(structs[struct_name])

    for f in fields:
        node = {
            "name": f["name"],
            "type": f["type"],
            "is_array": f["is_array"],
            "data_type": f["data_type"],
            "key": f["key"],
            "array_size": f["array_size"],
            "meta": f["meta"]
        }

        if f["type"] not in PRIMITIVES and f["type"] in structs:
            node["primitive"] = False
            node["children"] = build_tree_with_meta(f["type"], structs, visited.copy())
        else:
            node["primitive"] = True

        tree["fields"].append(node)

    return tree

# ---------------- Export Excel nâng cấp ----------------
def export_to_excel_with_meta(tree, output_file):
    """
    Export tree struct ra Excel, bao gồm các meta và array
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Struct Tree"

    # Header
    ws.append(["parent", "child_type", "child_name", "data_type", "key", "is_array","array_size","length_ref", "present", "bitmask", "range_check", "min", "max", "description"])

    # Flatten tree
    def flatten_tree(tree, rows=None):
        if rows is None:
            rows = []
        
        parent = tree.get("type")
        
        if parent in HANDLED_STRUCTS:
            return rows
        HANDLED_STRUCTS.add(parent)
        
        for field in tree.get("fields", []):
            meta = field.get("meta", {})

            rows.append([
                parent,
                field.get("type"),
                field.get("name"),
                field.get("data_type", ""),
                field.get("key", ""),
                field.get("is_array", False),
                field.get("array_size", ""),
                meta.get("length_ref", ""),
                meta.get("present", ""),
                meta.get("bitmask", ""),
                meta.get("range_check", ""),
                meta.get("min", ""),
                meta.get("max", ""),
                meta.get("description", "")
            ])
                
            if not field.get("primitive") and "children" in field:
                flatten_tree(field["children"], rows)
        return rows

    rows = flatten_tree(tree)
    rows.sort(key=lambda r: r[0])
    for r in rows:
        ws.append(r)

    format_excel(ws)

    wb.save(output_file)
    print(f" Exported to {output_file}")

# ---------------- Extract structs from code ----------------
def extract_structs(code):
    """
    Trích xuất tất cả typedef struct { ... } tên_struct;
    """
    structs = {}
    #struct_blocks = re.findall(r"typedef\s+struct\s*{(.*?)}\s*(\w+)\s*;", code, re.DOTALL)
    struct_blocks = re.findall(
        r"typedef\s+struct\s+(?:\w+\s*)?\{(.*?)\}\s*(\w+)\s*;",
        code,
        re.DOTALL
    )
    for body, name in struct_blocks:
        structs[name] = body
    return structs


def clear_comments(code):
    """
    Xóa tất cả comment C/C++ trừ những comment đặc biệt /*^ ... ^*/ dùng để lưu meta.
    """
    # 1. Giữ lại các comment /*^ ... ^*/ → tạm thay bằng placeholder
    placeholders = []
    def placeholder_replacer(match):
        placeholders.append(match.group(0))
        return f"__PLACEHOLDER_{len(placeholders)-1}__"

    code = re.sub(r"/\*\^(.*?)\^\*/", placeholder_replacer, code, flags=re.DOTALL)

    # 2. Xóa comment bình thường /* ... */ và // ...
    code = re.sub(r"/\*.*?\*/", "", code, flags=re.DOTALL)
    code = re.sub(r"//.*", "", code)

    # 3. Đưa lại comment đặc biệt
    for i, ph in enumerate(placeholders):
        code = code.replace(f"__PLACEHOLDER_{i}__", ph)

    return code

def fill_missing_data_in_excel(file_path):
    """
    Fill missing length_ref trong Excel:
    - Nếu range_check=OCTET_STRING và description=VARIABLE
    - Lấy 'child_name' của field khác cùng parent nếu parent có <= 2 fields
    """
    wb = load_workbook(file_path)
    ws = wb["Struct Tree"]

    # Lấy header để biết index của các cột
    headers = [cell.value for cell in ws[1]]
    idx_parent = headers.index("parent")
    idx_child_name = headers.index("child_name")
    idx_length_ref = headers.index("length_ref")
    idx_range_check = headers.index("range_check")
    idx_description = headers.index("description")

    # Tạo dict: parent -> list of rows (row objects)
    parent_map = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        parent = row[idx_parent].value
        parent_map.setdefault(parent, []).append(row)

    # Duyệt từng parent
    
    for parent, rows in parent_map.items():
        # Tìm các field cần fill
        for row in rows:
            range_check = row[idx_range_check].value
            description = row[idx_description].value
            length_ref = row[idx_length_ref].value
            child_name = row[idx_child_name].value

            if range_check == "OCTET_STRING" and description == "VARIABLE" and not length_ref:
                # Lấy các child_name khác cùng parent
                other_names = [r[idx_child_name].value for r in rows if r[idx_child_name].value != child_name]

                if len(other_names) == 0:
                    # chỉ có 1 field → không thể fill, skip
                    continue
                elif len(other_names) > 1:
                    # parent có >2 child → không xác định → raise error
                    raise ValueError(
                        f"Parent '{parent}' has more than 2 child fields, cannot determine length_ref for '{child_name}'"
                    )
                else:
                    # parent có đúng 2 child → fill length_ref
                    row[idx_length_ref].value = other_names[0]

    wb.save(file_path)
    print(f"[INFO] Filled missing length_ref in Excel: {file_path}")


#===============================================

def main():
    #code = read_file(DATA_C_PATH) + "\n" + read_file(OTHER_C_PATH)
    code = ""
    for filename in os.listdir(os.path.join(BASE_DIR, "..", "struct_input")):
        if filename.endswith(".c") or filename.endswith(".h"):
            code += "\n" + read_file(os.path.join(BASE_DIR, "..", "struct_input", filename))

    for filename in os.listdir(os.path.join(BASE_DIR, "..", "code_5G_header")):
        if filename.endswith(".c") or filename.endswith(".h"):
            code += "\n" + read_file(os.path.join(BASE_DIR, "..", "code_5G_header", filename))
    
    append_to_primitive_list(code)
    structs = extract_structs(code)
    tree = build_tree_with_meta(ROOT_STRUCT, structs)
    print(json.dumps(tree, indent=2))
    export_to_excel_with_meta(tree, DATA_EXCEL_PATH)
    fill_missing_data_in_excel(DATA_EXCEL_PATH)
    for key, types in PRIMITIVE_LIST.items():
        #print(f"{key:4} => {', '.join(types)}")
        print_log_info(f"{key:4} => {', '.join(types)}")

if __name__ == "__main__":
    main()